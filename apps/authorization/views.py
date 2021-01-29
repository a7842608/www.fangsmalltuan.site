import json
import re
import uuid
import random
import os
import datetime

import openpyxl
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q
from django.http import JsonResponse
from django.views import View
from django.contrib.auth import login, authenticate
from django_redis import get_redis_connection

from authorization.models import Users, MiddlePeople, UserLotteryNumber, UserCodeNumber, QuestionFeedback, MyPhoneCall, \
    MyWeChat, ExclusiveCustmer, IntegralSubsidiary, GoldenMoney, WhoLookMiddelPeople, GaoIndexPKMoney, \
    GaoBuildingDetialPKMoney, ReallyNameMiddlePeopleShare, AttentionMiddlePeople, UserHistoryMessageRecord, \
    AttentionVillage, WhoZanMiddelPeople, AttentionMiddlePeopleVillage, UserReport, AttentionAretical, UserSealStatus, \
    MessageChatListView, MiddleUnionId, UserEveryDayPermetionStatue
from index.models import BuildingDetial, LandDistrict, Article, BuildingImage, HouseImage, OneHouseOnePrice, Share, \
    Answer, Question, ZanCount
from .signals import middlepeople_viewed
from utils.auth import c2s, yixiangdengjibiao
from utils.jwt import token_check


def test_session(request):
    request.session['message'] = 'test django session OK'
    return JsonResponse({'Statue': 'success', 'Msg': '200'}, safe=False)


def __authorize_by_code(request):
    post = request.body.decode()
    uid = str(uuid.uuid4())
    suid = ''.join(uid.split('-'))
    post_str = post.split('=')
    code = post_str[1]
    # app_id ='wx3e2f018123732318'
    app_id = 'wx49eed93625eeb3ff'
    if not code:
        return JsonResponse({'Error': 'false', 'Msg': '接收信息不完整'})
    data = c2s(app_id, code)
    open_id = data.get('openid')
    session_key = data.get('session_key')
    if not open_id:
        return JsonResponse({'Error': 'false', 'Msg': 'openid错误'})

    request.session['openid'] = open_id
    request.session['session_key'] = session_key
    request.session['uuid'] = suid
    request.session['is_authorized'] = True

    # if not Users.objects.filter(open_id=open_id):
    #     new_user = Users(open_id=open_id, user_uuid=suid)
    #     new_user.save()
    return JsonResponse({"statue": 'success', 'open_id': open_id, 'session': request.session.get('session_key')},
                        safe=False)


class SafeUserWechatValueView(View):
    '''判断用户是否为首次登陆'''

    def get(self, request):
        try:
            # pan = request.GET.get('pan') # 0,1
            dat = request.GET.get('data')  # 手机号或open_id
            # if pan == '0':
            #     d = Users.objects.filter(mobile=dat).values()
            #     dl = list(d)
            #     if dl == []:
            #         return JsonResponse({'Error': 'false', 'Msg': '用户为未登录过用户, 没有手机号, 请绑定'})
            #     else:
            #         return JsonResponse({'Error': 'false', 'Msg': dl})
            # elif pan == '1':
            d = Users.objects.filter(open_id=dat).values()
            dl = list(d)
            if dl == []:
                return JsonResponse({'Error': 'false', 'Msg': True})
            else:
                return JsonResponse({'Error': 'false', 'Msg': dl})
            # else:
            #     return JsonResponse({'Error': 'false', 'Msg': '参数有误'})
        except Exception as e:
            return JsonResponse({'Error': 'false', 'Msg': e})

    '''存储用户img, nick_name, 地址'''

    def post(self, request):
        try:
            # mobile = request.POST.get('mobile')
            open_id = request.POST.get('open_id')
            if not open_id:
                return JsonResponse({'Error': 'false', 'Msg': 'openid错误'})
            nick_name = request.POST.get('nick_name')
            img = request.POST.get('img')
            adreess = request.POST.get('adreess')
            # sms_code = request.POST.get('sms_code')
            # if not all([mobile, sms_code]):
            # return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})
            # if not re.match(r'^1[3-9]\d{9}$', mobile):
            #     return JsonResponse({'code': 400, 'errmsg': 'mobile格式有误'})
            # redis_conn = get_redis_connection('sms_code')
            # sms_code_server = redis_conn.get('sms_%s' % mobile)
            # if not sms_code_server:
            #     return JsonResponse({'code': 400, 'errmsg': '短信验证码过期'})
            # if sms_code != sms_code_server.decode():
            #     return JsonResponse({'code': 400, 'errmsg': '验证码有误'})
            # pan = request.POST.get('pan')
            uid = str(uuid.uuid4())
            suid = ''.join(uid.split('-'))
            # if pan == '0':
            #     uu = Users.objects.values().filter(mobile=mobile)
            #     uu_list = list(uu)
            #     if uu_list == []:
            #         user = Users.objects.create(mobile=mobile, user_uuid=suid, nick_name=nick_name, header_img=img, adreess=adreess, really_name=nick_name, open_id=open_id, chat_room=suid)
            #     else:
            #         user = Users.objects.get(mobile=mobile)
            #         user.nick_name = nick_name
            #         user.header_img = img
            #         user.adreess = adreess
            #         user.open_id = open_id
            #         user.save()
            # elif pan == '1':
            uu = Users.objects.values().filter(open_id=open_id)
            uu_list = list(uu)
            if uu_list == []:
                user = Users.objects.create(user_uuid=suid, nick_name=nick_name, header_img=img, adreess=adreess,
                                            really_name=nick_name, open_id=open_id, chat_room=suid)
            else:
                user = Users.objects.get(open_id=open_id)
                user.nick_name = nick_name
                user.header_img = img
                user.adreess = adreess
                user.open_id = open_id
                user.save()

            a = Users.objects.values().filter(open_id=open_id)
            aa = list(a)

            return JsonResponse({"statue": 'success', 'Msg': '修改成功', 'user_info': aa}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class PanMobileView(View):
    '''判断用户是否绑定手机号'''

    def get(self, request):
        try:
            dat = request.GET.get('user_id')  # 手机号或open_id
            d = Users.objects.filter(id=dat).values('mobile')
            dl = list(d)
            if dl == []:
                return JsonResponse({'Error': 'false', 'Msg': False})
            else:
                return JsonResponse({'Error': 'false', 'Msg': dl})
        except Exception as e:
            return JsonResponse({'Error': 'false', 'Msg': e})

    '''手机号绑定'''

    def post(self, request):
        try:
            ur = request.POST.get('user_id')
            mobile = request.POST.get('mobile')
            sms_code = request.POST.get('sms_code')
            if not all([mobile, sms_code]):
                return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})
            if not re.match(r'^1[3-9]\d{9}$', mobile):
                return JsonResponse({'code': 400, 'errmsg': 'mobile格式有误'})
            redis_conn = get_redis_connection('sms_code')
            sms_code_server = redis_conn.get('sms_%s' % mobile)
            if not sms_code_server:
                return JsonResponse({'code': 400, 'errmsg': '短信验证码过期'})
            if sms_code != sms_code_server.decode():
                return JsonResponse({'code': 400, 'errmsg': '验证码有误'})

            user = Users.objects.get(id=ur)
            user.mobile = mobile
            user.save()
            uul = Users.objects.filter(id=ur).values()
            uu_list = list(uul)
            return JsonResponse({"statue": 'success', 'Msg': '验证成功', 'user_info': uu_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


def authorize(request):
    return __authorize_by_code(request)


class LoginView(View):
    '''顾问信息查询'''

    def get(self, request):
        try:
            dat = request.GET.get('user_id')
            d = MiddlePeople.objects.filter(user_fk_id=dat).values()
            dl = list(d)
            if dl == []:
                return JsonResponse({'Error': 'false', 'Msg': False})
            else:
                return JsonResponse({'Error': 'false', 'Msg': dl})
        except Exception as e:
            return JsonResponse({'Error': 'false', 'Msg': e})

    '''HTML5页面登陆接口'''

    def post(self, request):
        try:
            mobile = request.POST.get('mobile')
            sms_code = request.POST.get('sms_code')
            new_uu = str(uuid.uuid4())
            suid = ''.join(new_uu.split('-'))
            nick_name = '用户_' + str(random.randint(0, 9999999)) + '_'
            if not all([mobile, sms_code]):
                return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})
            if not re.match(r'^1[3-9]\d{9}$', mobile):
                return JsonResponse({'code': 400, 'errmsg': 'mobile格式有误'})
            redis_conn = get_redis_connection('sms_code')
            sms_code_server = redis_conn.get('sms_%s' % mobile)
            if not sms_code_server:
                return JsonResponse({'code': 400, 'errmsg': '短信验证码过期'})
            if sms_code != sms_code_server.decode():
                return JsonResponse({'code': 400, 'errmsg': '验证码有误'})
            data = Users.objects.values().filter(mobile=mobile)
            data_list = list(data)
            if data_list == []:
                try:
                    user = Users.objects.create(mobile=mobile, user_uuid=suid, nick_name=nick_name, header_img='暂无',
                                                really_name='空', chat_room=suid)
                    user_query = Users.objects.values().filter(mobile=mobile)
                    user_list = list(user_query)
                    context = {'Code': 200, 'Statue': '登陆成功您创建了一个新用户', 'Data': user_list}
                except Exception as e:
                    return JsonResponse({'code': 400, 'errmsg': {e}})
            else:
                try:
                    user_query = Users.objects.values().filter(mobile=mobile)
                    user_list = list(user_query)

                    if user_list == []:
                        return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
                    return JsonResponse({"Code": '200', 'Data': user_list}, safe=False)
                except Exception as e:
                    return JsonResponse({'code': 400, 'errmsg': {e}})
            # 实现状态保持
            # login(request, user)
            return JsonResponse(context)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class WechatBindingView(View):
    '''微信绑定接口'''

    def post(self, request):
        try:
            uu = request.POST.get('uu')
            mo = request.POST.get('mo')
            sms_code = request.POST.get('sms_code')
            if not all([uu, mo, sms_code]):
                return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})
            if not re.match(r'^1[3-9]\d{9}$', mo):
                return JsonResponse({'code': 400, 'errmsg': 'mobile格式有误'})
            redis_conn = get_redis_connection('sms_code')
            sms_code_server = redis_conn.get('sms_%s' % mo)
            if not sms_code_server:
                return JsonResponse({'code': 400, 'errmsg': '短信验证码过期'})
            if sms_code != sms_code_server.decode():
                return JsonResponse({'code': 400, 'errmsg': '验证码有误'})
            # 查询数据库中用户修改
            data = Users.objects.get(user_uuid=uu)
            data.mobile = mo
            data.save()

            context = {'Code': 200, 'Statue': 'success', 'Data': '修改成功'}
            # 实现状态保持
            # login(request, user)
            return JsonResponse(context)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserCenterSystemView(View):
    '''个人中心三个数'''

    def get(self, request):
        try:
            req = request.GET.get('id')
            us_lo = UserLotteryNumber.objects.filter(user_id_id=req).count()  # 用户摇号数量
            us_ch = UserHistoryMessageRecord.objects.filter(fk_id=req).count()  # 聊天记录数量

            at_bd = AttentionVillage.objects.filter(user_id=req).count()
            at_mp = AttentionMiddlePeople.objects.filter(user_id=req).count()  # 关注数量
            at_atc = AttentionAretical.objects.filter(user_id=req).count()

            two = at_bd + at_mp + at_atc

            data_list = []
            json_dict = {}
            json_dict['lottery_number'] = us_lo
            json_dict['middle_people'] = two
            json_dict['message'] = us_ch
            data_list.append(json_dict)
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)

    '''获取用户信息'''

    def post(self, request):
        try:
            req = request.POST.get('open_id')
            data = Users.objects.filter(open_id=req).values()
            data_list = list(data)
            if data_list == []:
                return JsonResponse({"statue": '404', 'data': 'not_found_users'}, safe=False)
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)


class UsersAttionBuildingView(View):
    '''普通用户关注楼盘'''

    def get(self, request):
        try:
            req = request.GET.get('user_id')
            query_data = AttentionVillage.objects.filter(user_id=req).values('building_id')
            query_list = list(query_data)
            data_list = []
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400, 'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            for a in page_value:
                a_id = a['building_id']
                data_l = BuildingDetial.objects.filter(id=a_id).values()
                im = BuildingImage.objects.filter(id=a_id).values('photo_image')
                da = list(data_l)
                for i in da:
                    json_dict = {}
                    for t in im:
                        json_dict['img'] = t['photo_image']
                    json_dict['id'] = i['id']
                    json_dict['building_name'] = i['building_name']
                    json_dict['unit_price'] = i['unit_price']
                    aa = i['land_id']
                    query_other_data = LandDistrict.objects.get(id=aa)
                    json_dict['land_id'] = query_other_data.name
                    json_dict['comment_count'] = i['comment_count']
                    json_dict['sale_stage'] = i['sale_stage']
                    data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)


class UsersAttionArticalgView(View):
    '''普通用户关注文章'''

    def get(self, request):
        try:
            req = request.GET.get('user_id')
            query_data = AttentionAretical.objects.filter(user_id=req).values('aretical_id')
            query_list = list(query_data)
            data_list = []
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400, 'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            for a in page_value:
                a_id = a['aretical_id']
                data_l = Article.objects.filter(id=a_id).values()
                da = list(data_l)
                for i in da:
                    json_dict = {}
                    json_dict['choice_classfiy'] = i['choice_classfiy']
                    json_dict['id'] = i['id']
                    json_dict['author'] = i['author']
                    json_dict['author_img'] = i['author_img']
                    json_dict['create_time'] = i['create_time']
                    json_dict['title'] = i['title']
                    json_dict['content'] = i['content']
                    json_dict['new_img'] = i['new_img']
                    json_dict['land'] = i['land']
                    data_list.append(json_dict)
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)


class UsersAttionMiddlePeopleView(View):
    '''普通用户关注顾问'''

    def get(self, request):
        try:
            req = request.GET.get('user_id')
            query_data = AttentionMiddlePeople.objects.filter(user_id=req).values('middle_people_id')
            query_list = list(query_data)
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400, 'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for a in page_value:
                a_id = a['middle_people_id']
                dl = MiddlePeople.objects.filter(id=a_id).values()
                dl_list = list(dl)
                for i in dl_list:
                    json_dict = {}
                    json_dict['header_img'] = i['header_img']
                    json_dict['id'] = i['id']
                    u = Users.objects.get(id=i['id'])
                    json_dict['chat_room'] = u.chat_room
                    json_dict['nick_name'] = i['nick_name']
                    json_dict['rank'] = i['rank']
                    json_dict['browse_count'] = i['browse_count']
                    json_dict['live_limit'] = i['live_limit']
                    json_dict['click_count'] = i['click_count']
                    a = i['building_fk_id']
                    d = BuildingDetial.objects.values().get(id=a)
                    json_dict['building_name'] = d['building_name']
                    data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': str(e)}
            return JsonResponse(context)


class MyLotteryView(View):
    '''我的摇号'''

    def get(self, request):
        try:
            req = request.GET.get('id')  # 用户id
            data = UserLotteryNumber.objects.values().filter(user_id_id=req).order_by('-create_time')
            query_list = list(data)
            # 分页
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400,
                                     'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['create_time'] = i['create_time']
                json_dict['really_name'] = i['really_name']
                json_dict['number'] = i['number']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)

    # 我的摇号之添加新用户
    def post(self, request):
        try:
            user_id = request.POST.get('user_id')  # 5
            if not user_id:
                return JsonResponse({'Error': 'false', 'Msg': '用户错误'})
            really_name = request.POST.get('really_name')
            ID_card = request.POST.get('ID_card')
            user = Users.objects.get(id=user_id)
            user.really_name = really_name
            user.ID_card = ID_card
            user.save()
            UserLotteryNumber.objects.create(user_id_id=user_id, really_name=really_name, number=ID_card)
            return JsonResponse({"statue": 'success', 'Msg': '保存成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class BuildingDetialMyLotteryView(View):
    '''楼盘详情页我的摇号接口显示'''

    def get(self, request):
        try:
            req = request.GET.get('id')  # 用户id
            bd = request.GET.get('only_id')  # 楼盘id
            data = UserCodeNumber.objects.values().filter(Q(build_id_id=bd) & Q(user_id_id=req)).order_by(
                '-create_time')
            query_list = list(data)
            data_list = []
            for i in query_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['really_name'] = i['really_name']
                json_dict['number'] = i['number']
                json_dict['create_time'] = i['create_time']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)

    def post(self, request):
        try:
            req = request.POST.get('id')  # 用户id
            bd = request.POST.get('only_id')  # 楼盘id
            rel = request.POST.get('really_name')  # 楼盘id
            data = UserCodeNumber.objects.create(user_id_id=req, build_id_id=bd, really_name=rel)
            return JsonResponse({"statue": 'success', 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)


class ChoiceBuildingNameView(View):
    '''选择楼盘下拉框'''

    def get(self, request):
        try:
            data = BuildingDetial.objects.all().values('id')
            query_list = list(data)
            # 分页
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400,
                                     'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                a = i['id']
                d = BuildingDetial.objects.values().get(id=a)
                json_dict['building_name'] = d['building_name']
                json_dict['id'] = a
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)


class MyCodeView(View):
    '''我的编码'''

    def get(self, request):
        try:
            req = request.GET.get('id')  # 用户id
            data = UserCodeNumber.objects.values().filter(user_id__id=req)
            query_list = list(data)
            # 分页
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400,
                                     'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['create_time'] = i['create_time']
                json_dict['number'] = i['number']
                # 用户名
                a = i['user_id_id']
                aa = Users.objects.values().get(id=a)
                json_dict['really_name'] = aa['really_name']
                # 楼盘名
                b = i['build_id_id']
                bb = BuildingDetial.objects.values().get(id=b)
                json_dict['building_name'] = bb['building_name']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"RESULT": 'false', 'MSG': {e}}
            return JsonResponse(context)

    '''添加编码'''

    def post(self, request):
        try:
            user_id = request.POST.get('user_id')  # 5
            if not user_id:
                return JsonResponse({'Error': 'false', 'Msg': '用户错误'})
            build_id = request.POST.get('build_id')
            number = request.POST.get('number')
            really_name = request.POST.get('really_name')
            if not all([user_id, build_id, number, really_name]):
                return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})

            UserCodeNumber.objects.create(
                build_id_id=build_id,
                really_name=really_name,
                number=number,
                user_id_id=user_id
            )
            try:
                union = MiddleUnionId.objects.get(fk_id=user_id)
                bd = BuildingDetial.objects.get(id=build_id)
                ur = Users.objects.get(id=user_id)
                yixiangdengjibiao(union.gong_open_id, bd.building_name, really_name, ur.mobile)
            except:
                return JsonResponse({"statue": 'success', 'Msg': '添加编码成功, 但你没有关注本小程序公众号, 请关注公众号以便于进行消息提醒, 谢谢'},
                                    safe=False)

            return JsonResponse({"statue": 'success', 'Msg': '保存成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class DeleteMyLotteryNumberView(View):
    # 删除我的编码
    def post(self, request):
        try:
            sql_id = request.POST.get('id')
            # print(sql_id, type(sql_id))
            de = UserCodeNumber.objects.get(id=sql_id)
            de.delete()
            return JsonResponse({"statue": 'success', 'Msg': '删除成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class QuestionUploadView(View):
    '''问题反馈'''

    def post(self, request):
        try:
            choice = request.POST.get('choice')  # 0,1,2,3
            content = request.POST.get('content')
            phone = request.POST.get('phone')
            img = request.POST.get('img')
            img1 = request.POST.get('img1')

            user = request.POST.get('user_id')

            QuestionFeedback.objects.values().create(user_id=user, choice_classfiy=choice, my_content=content,
                                                     feedback_phone=phone, question_img=img, img1=img1)

            return JsonResponse({"statue": 'success', 'Msg': '保存成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddlePeopleGetView(View):
    '''置业顾问入住'''

    def post(self, request):
        try:
            aa_id = request.POST.get('user_id')  # 用户id
            building_id = request.POST.get('building_id')  # 楼盘id
            nick_name = request.POST.get('nick_name')  # 昵称
            mobile = request.POST.get('mobile')  # 手机号
            header_img = request.POST.get('header_img')  # 头像
            name = request.POST.get('name')  # 真实用户名称
            wx_number = request.POST.get('wx_number')  # 微信号
            choice_building = request.POST.get('choice_building')  # 选择主营楼盘
            please_number = request.POST.get('please_number')  # 邀请码
            wx_two = request.POST.get('wx_two')  # 二维码
            work_pai = request.POST.get('work_pai')  # 工牌

            people = request.POST.get('if_mp')

            sms_code = request.POST.get('sms_code')  # 短信验证码

            if not all([mobile, sms_code]):
                return JsonResponse({'code': 400, 'errmsg': '缺少必传参数'})
            if not re.match(r'^1[3-9]\d{9}$', mobile):
                return JsonResponse({'code': 400, 'errmsg': 'mobile格式有误'})
            redis_conn = get_redis_connection('sms_code')
            sms_code_server = redis_conn.get('sms_%s' % mobile)
            if not sms_code_server:
                return JsonResponse({'code': 400, 'errmsg': '短信验证码过期'})
            if sms_code != sms_code_server.decode():
                return JsonResponse({'code': 400, 'errmsg': '验证码有误'})

            user = Users.objects.get(id=aa_id)
            user.if_middle_people = '1'
            user.save()
            # print(user)

            query_data = MiddlePeople.objects.create(
                user_fk_id=aa_id,
                building_fk_id=building_id,
                nick_name=nick_name,
                mobile=mobile,
                header_img=header_img,
                really_name=name,
                wechat_number=wx_number,
                bussiness_building=choice_building,
                invitation_code=please_number,
                two_wei_ma=wx_two,
                work_pai=work_pai
            )

            iid = MiddlePeople.objects.values().get(mobile=mobile)
            ii_d = iid['id']

            AttentionMiddlePeopleVillage.objects.create(building_id=building_id, user_id=ii_d)

            Users.objects.filter(id=aa_id).update(middle_id=ii_d)

            return JsonResponse({"statue": 'success', 'data': '成功变成顾问', 'PeopleStatue': '1', 'mp_id': ii_d}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


'''顾问之家'''


class MiddlePeopleHouseDetialView(View):
    '''顾问之家首页'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MiddlePeople.objects.values().filter(id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['header_img'] = i['header_img']
                json_dict['really_name'] = i['really_name']
                json_dict['rank'] = i['rank']
                json_dict['call_mobile'] = i['call_mobile']
                json_dict['wechat_talk'] = i['wechat_talk']
                json_dict['exclusive_people'] = i['exclusive_people']
                json_dict['integral'] = i['integral']
                json_dict['live_limit'] = i['live_limit']
                json_dict['golden_money'] = i['golden_money']
                json_dict['user_fk_id'] = i['user_fk_id']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddlePhoneView(View):
    '''顾问之家电话'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MyPhoneCall.objects.values().filter(middle_people__id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['middle_people_id'] = i['middle_people_id']
                json_dict['head_img'] = i['head_img']
                json_dict['name'] = i['name']
                json_dict['user_phone'] = i['user_phone']
                json_dict['phone_time'] = i['phone_time']
                json_dict['choice_classfiy'] = i['choice_classfiy']
                json_dict['user_id'] = i['user_id']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MyWeChatView(View):
    '''顾问之家聊天'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MyWeChat.objects.values().filter(middle_name__id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['middle_name_id'] = i['middle_name_id']
                json_dict['head_img'] = i['head_img']
                json_dict['user_name'] = i['user_name']
                json_dict['where_come_from'] = i['where_come_from']
                json_dict['custmer_time'] = i['custmer_time']
                json_dict['choice_classfiy'] = i['choice_classfiy']
                json_dict['user_id'] = i['user_id']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddleCustmerView(View):
    '''顾问之家专属客户'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = ExclusiveCustmer.objects.values().filter(middle_name_id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['middle_name_id'] = i['middle_name_id']
                json_dict['head_img'] = i['head_img']
                json_dict['user_name'] = i['user_name']
                json_dict['custmer_time'] = i['custmer_time']
                json_dict['user_id'] = i['user_id']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddleFenView(View):
    '''顾问之家积分明细表'''

    def get(self, request):
        try:
            # tk = request.META['HTTP_TOKEN']
            # atk = token_check(tk)
            # if atk != 'True':
            #     return JsonResponse({'Error': 'false', 'Msg': 'Token_Fales'})

            mp = request.GET.get('mp_id')
            ye = request.GET.get('year')  # 2020
            mo = request.GET.get('month')  # 12
            da = int(request.GET.get('day'))  # 30

            a = MiddlePeople.objects.get(id=mp)
            j = {'integral': a.integral}
            data_list = []
            data_list.append(j)
            for i in range(1, da):
                jd = {}
                tim = ye + '-' + mo + '-' + str(i)
                ti = datetime.datetime.strptime(tim, '%Y-%m-%d').date()
                jd['day'] = ti
                req = IntegralSubsidiary.objects.filter(Q(fk_id=mp) & Q(score_create_time=ti)).values()
                req_list = list(req)
                if req_list == []:
                    continue
                else:
                    dl = []
                    for i in req_list:
                        json_dict = {}
                        json_dict['id'] = i['id']
                        json_dict['fk_id'] = i['fk_id']
                        json_dict['score'] = i['score']
                        json_dict['score_create_time'] = i['score_create_time']
                        json_dict['change_beacuse'] = i['change_beacuse']
                        json_dict['create_time'] = i['create_time']
                        json_dict['choice_classfiy'] = i['choice_classfiy']
                        dl.append(json_dict)
                    jd['son'] = dl
                    data_list.append(jd)
            if data_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': str(e)}
            return JsonResponse(context)

    '''判断用户是否签到了今天'''

    def post(self, request):
        try:
            mp = request.POST.get('mp_id')
            import datetime
            date = datetime.datetime.now().date()
            ur = MiddlePeople.objects.get(id=mp)
            ur_id = ur.id
            dayl = IntegralSubsidiary.objects.filter(
                Q(fk_id=mp) & Q(create_time=date) & Q(change_beacuse='每日签到')).values()
            day = list(dayl)
            if day == []:
                try:
                    union = MiddleUnionId.objects.get(fk_id=ur_id)
                    jifenduihuantongzhi(union.gong_open_id)
                except:
                    return JsonResponse({'Statue': 'false', 'Msg': '用户未关注公众号, 需要关注公众号才能够发送推送消息提示其未签到'})
                return JsonResponse({'Statue': 200, 'Msg': False})
            else:
                return JsonResponse({"statue": 200, 'data': True}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddleMoneyView(View):
    '''顾问之家金币明细表'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = GoldenMoney.objects.values().filter(fk_id=mp)
            req_list = list(req)
            # 分页
            paginator = Paginator(req_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400, 'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                json_dict['id'] = i['id']  # id
                json_dict['money_count'] = i['money_count']  # 修改数量
                json_dict['recharge_time'] = i['recharge_time']  # 修改时间
                json_dict['change_beacuse'] = i['change_beacuse']  # 修改原因
                json_dict['create_time'] = i['create_time']  # 数据框见时间
                json_dict['choice_classfiy'] = i['choice_classfiy']  # 0收入, 1指出
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class SeaPaperView(View):
    '''专属海报'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MiddlePeople.objects.values().filter(id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['header_img'] = i['header_img']
                json_dict['really_name'] = i['really_name']
                json_dict['rank'] = i['rank']
                json_dict['click_count'] = i['click_count']
                json_dict['integral'] = i['integral']
                json_dict['browse_count'] = i['browse_count']
                json_dict['invitation_code'] = i['invitation_code']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ShowPageView(View):
    '''展业工具'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MiddlePeople.objects.filter(id=mp).values('id', 'building_fk_id')
            build = req[0]['building_fk_id']
            query_data = BuildingImage.objects.filter(Q(choice_classfiy=1) & Q(fk_id=build)).values('photo_image')
            img = query_data[0]['photo_image']
            name = BuildingDetial.objects.filter(id=build).values('id', 'building_name', 'premises_location')
            data_list = []
            json_dict = {}
            json_dict['img'] = img
            json_dict['building_id'] = name[0]['id']
            json_dict['building_name'] = name[0]['building_name']
            json_dict['premises_location'] = name[0]['premises_location']
            data_list.append(json_dict)
            if query_data == []:
                return JsonResponse({'Statue': 'false', 'Msg': '没有图片'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''展业工具之楼盘排名'''

    def post(self, request):
        try:
            mp = request.POST.get('building_fk_id')
            req = BuildingDetial.objects.all().values('id')
            build = list(req)
            data_list = []
            for a, b in enumerate(build):
                v = b['id']
                json_dict = {}
                if str(v) == mp:
                    print(v)
                    json_dict['index'] = str(a + 1)
                    data_list.append(json_dict)
                else:
                    continue
            if data_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddleVillageView(View):
    '''我的楼盘(他入住的楼盘)'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = AttentionMiddlePeopleVillage.objects.filter(user_id=mp).values('building_id')
            try:
                build_id = req[0]['building_id']
            except Exception as e:
                return JsonResponse({'Statue': 'false', 'Msg': '你没有楼盘,查锤子'})
            q = BuildingDetial.objects.values().filter(id=build_id)
            q_l = list(q)
            paginator = Paginator(q_l, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400, 'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['building_name'] = i['building_name']
                json_dict['premises_location'] = i['premises_location']
                json_dict['if_'] = '我是从置业顾问入住楼盘进来的'
                a = i['id']
                img = BuildingImage.objects.filter(fk_id=a).values('photo_image')
                img_list = list(img)
                for t in img_list:
                    json_dict['img'] = t['photo_image']
                data_list.append(json_dict)
            if build_id == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MyBuildingStatueView(View):
    '''我的楼盘之楼盘状态'''

    def get(self, request):
        try:
            bp = request.GET.get('only_id')

            req = BuildingStatueTimeSale.objects.values().filter(fk_id=bp)
            req_list = list(req)
            req2 = BuildingDetial.objects.filter(id=bp).values('sale_stage', 'lottery_count', 'building_create_time')

            data1_list = []
            for b in req2:
                json_dict1 = {}
                json_dict1['sale_stage'] = b['sale_stage']
                json_dict1['lottery_count'] = b['lottery_count']
                json_dict1['building_create_time'] = b['building_create_time']
                data1_list.append(json_dict1)

            if data1_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})

            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['fk_id'] = i['fk_id']
                json_dict['will_sale_time'] = i['will_sale_time']
                json_dict['register_time'] = i['register_time']
                json_dict['commit_time'] = i['commit_time']
                json_dict['want_told_time'] = i['want_told_time']
                json_dict['lottery_time'] = i['lottery_time']
                json_dict['choice_house_time'] = i['choice_house_time']

                data_list.append(json_dict)
            if data_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'data1': data1_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class CreateHouseXingImgView(View):
    '''生成户型海报'''

    def get(self, request):
        try:
            mp = request.GET.get('only_id')  # 楼盘详情页
            req = HouseImage.objects.values().filter(building_detial__id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['image'] = i['image']
                json_dict['create_time'] = i['create_time']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class CreateMingPianImgView(View):
    '''生成名片海报'''

    def get(self, request):
        try:
            mp = request.GET.get('middle_id')  # 楼盘详情页
            # bp = request.GET.get('only_id') # 楼盘详情页
            req = MiddlePeople.objects.values().filter(id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                a = i['building_fk_id']
                aa = BuildingDetial.objects.values().get(id=a)
                json_dict['id'] = i['id']
                json_dict['building_name'] = aa['building_name']
                json_dict['wechat_number'] = i['wechat_number']
                json_dict['mobile'] = i['mobile']
                json_dict['two_wei_ma'] = i['two_wei_ma']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddlePeopleOneHouseOnePriceCreateView(View):
    '''生成一房一价'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = MiddlePeople.objects.values('two_wei_ma', 'building_fk_id', 'header_img', 'really_name',
                                              'mobile').filter(id=mp)
            req_list = list(req)
            for i1 in req_list:
                i = i1['building_fk_id']
                ii = i1['header_img']
                iii = i1['really_name']
                iiii = i1['mobile']
                two = i1['two_wei_ma']
                data_list = []
                bd = BuildingDetial.objects.get(id=i)
                ohop = OneHouseOnePrice.objects.filter(building_detial_id=i).values('house_dong', 'house_yuan',
                                                                                    'door_number', 'one_price',
                                                                                    'all_price')
                ohop_list = list(ohop)
                for a in ohop_list:
                    json_dict = {}
                    json_dict['building_name'] = bd.building_name
                    json_dict['only_id'] = bd.id
                    json_dict['premises_location'] = bd.premises_location
                    json_dict['house_dong'] = a['house_dong']
                    json_dict['house_yuan'] = a['house_yuan']
                    # json_dict['house_ceng'] = a['house_ceng']
                    json_dict['door_number'] = a['door_number']
                    json_dict['one_price'] = a['one_price']
                    json_dict['all_price'] = a['all_price']
                    json_dict['header_img'] = ii
                    json_dict['really_name'] = iii
                    json_dict['mobile'] = iiii
                    json_dict['two_wei_ma'] = two
                    data_list.append(json_dict)
                if req_list == []:
                    return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
                return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    # 下拉菜单
    def post(self, request):
        try:
            bd = request.POST.get('only_id')
            dong = request.POST.get('dong')
            bdid = BuildingDetial.objects.get(id=bd)
            ohop = OneHouseOnePrice.objects.filter(Q(building_detial_id=bd) & Q(house_dong=dong)).values('house_dong',
                                                                                                         'house_yuan',
                                                                                                         'house_ceng',
                                                                                                         'door_number',
                                                                                                         'one_price',
                                                                                                         'all_price')
            ohop_list = list(ohop)
            data_list = []
            for a in ohop_list:
                json_dict = {}
                json_dict['building_name'] = bdid.building_name
                json_dict['only_id'] = bdid.id
                json_dict['premises_location'] = bdid.premises_location
                json_dict['house_dong'] = a['house_dong']
                json_dict['house_yuan'] = a['house_yuan']
                json_dict['house_ceng'] = a['house_ceng']
                json_dict['door_number'] = a['door_number']
                json_dict['one_price'] = a['one_price']
                json_dict['all_price'] = a['all_price']
                data_list.append(json_dict)
            if ohop_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MiddlePeopleDetialView(View):
    view_signal = middlepeople_viewed

    '''置业顾问详情页'''

    def get(self, request):

        try:
            mp = request.GET.get('mp_id')
            middlepeople = MiddlePeople.objects.get(pk=int(mp))
            self.send_signal(request, middlepeople)

            req = MiddlePeople.objects.values().filter(id=mp)
            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['user_fk_id'] = i['user_fk_id']
                json_dict['img'] = i['header_img']
                json_dict['really_name'] = i['really_name']
                json_dict['rank'] = i['rank']
                json_dict['wechat_number'] = i['wechat_number']
                json_dict['bussiness_building'] = i['bussiness_building']
                json_dict['browse_count'] = i['browse_count']
                json_dict['live_limit'] = i['live_limit']
                json_dict['click_count'] = i['click_count']
                json_dict['two_wei_ma'] = i['two_wei_ma']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def post(self, request):
        '''帮我点赞(加赞)'''
        try:
            middle_id = request.POST.get('mp_id')
            zan1 = request.POST.get('zan')
            user_id = request.POST.get('user_id')

            cou = WhoZanMiddelPeople.objects.filter(Q(middle_id=middle_id) & Q(user_id=user_id)).count()
            if cou == 0:
                ZanCount.objects.create(user_id=user_id, type_id=middle_id, choice_classfiy=3)
                WhoZanMiddelPeople.objects.create(user_id=user_id, middle_id=middle_id, fk_id=middle_id)
                query_data = MiddlePeople.objects.values().get(id=middle_id)
                allcount = query_data['click_count']
                zan = 1 + int(allcount)
                MiddlePeople.objects.filter(id=middle_id).update(click_count=zan)
                return JsonResponse({"statue": 'success', 'data': '点赞成功'}, safe=False)
            else:
                return JsonResponse({"statue": 'false', 'data': '已经点过了'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def send_signal(self, request, middlepeople):
        self.view_signal.send(
            sender=self, middlepeople=middlepeople, user=request.user, request=request, )


class WhoLookMeView(View):
    '''浏览量/谁看过我'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            req = WhoLookMiddelPeople.objects.values().filter(fk_id=mp)
            req_list = list(req)
            # print(req_list)
            data_list = []
            for i in req_list:
                json_dict = {}
                a = i['user_id']
                data = Users.objects.values().get(id=a)
                json_dict['header_img'] = data['header_img']
                json_dict['nick_name'] = data['nick_name']
                data_list.append(json_dict)
            if req_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)

        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def post(self, request):
        # 添加访客记录
        try:
            mp_id = request.POST.get('mp_id')
            user_id = request.POST.get('user_id')
            u = WhoLookMiddelPeople.objects.filter(Q(user_id=user_id) & Q(fk_id=mp_id)).count()
            if u >= 1:
                # f = WhoLookMiddelPeople.objects.values('id').filter(Q(fk_id=mp_id)&Q(user_id=user_id))
                # f_l = list(f)
                # for inde, a in enumerate(f_l):
                #     if inde >= 1:
                return JsonResponse({"statue": 'success', 'data': '我过来过了,还不止一次'}, safe=False)
                #     t = WhoLookMiddelPeople.objects.get(id=a)
                #     t.user_id=user_id
                #     t.save()
                # return JsonResponse({"statue": 'success', 'data': '我过来过了,还不止一次'}, safe=False)
            else:
                WhoLookMiddelPeople.objects.create(
                    fk_id=mp_id, user_id=user_id
                )
                return JsonResponse({"statue": 'success', 'data': '我过来过了,就一回'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class HeFenXiangView(View):
    '''他的分享(分享堂)'''

    def get(self, request):
        try:
            data = request.GET.get('mp_id')  # 顾问名称
            query_data = Share.objects.values().filter(middle_fk_id=data)
            query_list = list(query_data)
            # 分页
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400,
                                     'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                json_dict['id'] = i['id']
                a = MiddlePeople.objects.values().get(id=data)
                json_dict['header_img'] = a['header_img']
                json_dict['building_fk_id'] = i['building_fk_id']
                json_dict['really_name'] = a['really_name']
                json_dict['building_name'] = i['building_name']
                json_dict['create_time'] = i['create_time']
                json_dict['content'] = i['content']
                json_dict['img'] = i['img']
                json_dict['video'] = i['video']
                json_dict['choice_classfiy'] = i['choice_classfiy']
                json_dict['browse_count'] = i['browse_count']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''上传分享'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            Share.objects.create(**json_dict)
            return JsonResponse({"statue": 200, 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class HeAnswerView(View):
    '''他的回答(问)'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问名称
            query_data = Answer.objects.values().filter(aut_id=mp)
            query_list = list(query_data)
            # 分页
            paginator = Paginator(query_list, 10)
            page = int(request.GET.get('page'))
            try:
                page_value = paginator.page(page)
            except EmptyPage:
                return JsonResponse({'Code': 400,
                                     'Errmsg': 'page数据出错'})
            total_page = paginator.num_pages
            data_list = []
            for i in page_value:
                json_dict = {}
                a = i['question_id']
                json_dict['id'] = i['id']
                # try:
                aa = Question.objects.values().filter(id=a)
                qd = list(aa)
                for aaa in qd:
                    json_dict['title'] = aaa['title']
                # except:
                # return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
                json_dict['content'] = i['content']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''给顾问取消赞'''

    def post(self, request):
        try:
            user_id = request.POST.get('user_id')
            type_id = request.POST.get('mp_id')
            try:
                ZanCount.objects.filter(Q(user_id=user_id) & Q(type_id=type_id) & Q(choice_classfiy=3)).delete()
                WhoZanMiddelPeople.objects.filter(Q(user_id=user_id) & Q(middle_id=type_id) & Q(fk_id=type_id)).delete()
            except:
                return JsonResponse({"statue": 'false', 'msg': '取消点赞成功'}, safe=False)
            return JsonResponse({"statue": 'success', 'msg': '成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ClickMiddlePeopleAttentionView(View):
    '''取消关注置业顾问'''

    def get(self, request):
        try:
            mp_id = request.GET.get('mp_id')  # 顾问id
            user_id = request.GET.get('user_id')  # 用户id
            # query_data = ExclusiveCustmer.objects.values.filter(middle_name_id=mp_id)
            try:
                exclus = ExclusiveCustmer.objects.get(Q(middle_name__id=mp_id) & Q(user_id=user_id))
                exclus.delete()
                # z = query_data.click_count
                # c_z = int(z) - 1
                # query_data.click_count = c_z
                # query_data.save()
            except:
                return JsonResponse({"statue": 'false', 'msg': '删除失败'}, safe=False)
            return JsonResponse({"statue": 'success', 'msg': '取消成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''点击关注顾问'''

    def post(self, request):
        try:
            mp = request.POST.get('mp_id')  # 顾问ID
            at = request.POST.get('user_id')
            user_name = request.POST.get('user_name')
            # custmer_time = request.POST.get('custmer_time')
            head_img = request.POST.get('head_img')
            # query_data = ExclusiveCustmer.objects.create(middle_name_id=mp, user_id=at, user_name=user_name, custmer_time=custmer_time, head_img=head_img)
            query_data = ExclusiveCustmer.objects.create(middle_name_id=mp, user_id=at, user_name=user_name,
                                                         head_img=head_img)
            return JsonResponse({"statue": 'success', 'data': '关注成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class MyWillGoView(View):
    '''全站/楼盘竞价页'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            query_data = MiddlePeople.objects.values().filter(id=mp)
            query_list = list(query_data)
            data_list = []
            for i in query_list:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['golden_money'] = i['golden_money']
                a = GaoIndexPKMoney.objects.values().all().order_by('-create_time')
                aa = list(a)
                # for ax in aa:
                json_dict['index_price'] = aa[0]['price']
                json_dict['fk_id'] = aa[0]['fk_id']
                # json_dict['fk_id'] = ax['fk_id']
                b = GaoBuildingDetialPKMoney.objects.values().all().order_by('-create_time')
                bb = list(b)
                # for bx in bb:
                json_dict['detial_price'] = bb[0]['price']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class PKMoneyView(View):
    '''顾问竞价详情页'''

    def get(self, request):
        try:
            if request.GET.get('choice') == '1':  # 首页出价
                data = request.GET.get('mp_id')  # 顾问名称
                query_data = GaoIndexPKMoney.objects.values().filter(fk_id=data).order_by('-create_time')
                query_list = list(query_data)
                # 分页
                paginator = Paginator(query_list, 10)
                page = int(request.GET.get('page'))
                try:
                    page_value = paginator.page(page)
                except EmptyPage:
                    return JsonResponse({'Code': 400,
                                         'Errmsg': 'page数据出错'})
                total_page = paginator.num_pages
                data_list = []
                for i in page_value:
                    json_dict = {}
                    json_dict['id'] = i['id']
                    json_dict['building_id'] = i['building_id']
                    json_dict['price'] = i['price']
                    json_dict['create_time'] = i['create_time']
                    data_list.append(json_dict)
                if query_list == []:
                    return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
                return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)

            elif request.GET.get('choice') == '2':  # 楼盘详情页出价
                data = request.GET.get('mp_id')  # 顾问名称
                query_data = GaoBuildingDetialPKMoney.objects.values().filter(fk_id=data).order_by('-create_time')
                query_list = list(query_data)
                # 分页
                paginator = Paginator(query_list, 10)
                page = int(request.GET.get('page'))
                try:
                    page_value = paginator.page(page)
                except EmptyPage:
                    return JsonResponse({'Code': 400,
                                         'Errmsg': 'page数据出错'})
                total_page = paginator.num_pages
                data_list = []
                for i in page_value:
                    json_dict = {}
                    json_dict['id'] = i['id']
                    json_dict['building_id'] = i['building_id']
                    json_dict['price'] = i['price']
                    json_dict['create_time'] = i['create_time']
                    data_list.append(json_dict)
                if query_list == []:
                    return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
                return JsonResponse({"statue": 'success', 'data': data_list, 'Count': total_page}, safe=False)
            else:
                context = {"Result": 'false', 'Msg': '参数错误'}
                return JsonResponse(context)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''出价'''

    def post(self, request):
        try:
            if request.POST.get('choice') == '1':  # 首页
                fk_id = request.POST.get('fk_id')  # 顾问ID
                # money = request.POST.get('money')
                building_id = request.POST.get('building_id')
                price = request.POST.get('price')
                # head_img = request.POST.get('head_img')
                query_data = GaoIndexPKMoney.objects.create(fk_id=fk_id, building_id=building_id, price=price)
                return JsonResponse({"statue": 'success', 'data': '关注成功'}, safe=False)

            elif request.POST.get('choice') == '2':  # 楼盘详情
                fk_id = request.POST.get('fk_id')  # 顾问ID
                # money = request.POST.get('money')
                building_id = request.POST.get('building_id')
                price = request.POST.get('price')
                # head_img = request.POST.get('head_img')
                query_data = GaoBuildingDetialPKMoney.objects.create(fk_id=fk_id, building_id=building_id, price=price)
                return JsonResponse({"statue": 'success', 'data': '关注成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class RequestGlodenMoneyConversionView(View):
    '''获取金币兑换页'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            query_data = MiddlePeople.objects.filter(id=mp).values('id', 'golden_money', 'integral')
            data_list = []
            for i in query_data:
                json_dict = {}
                json_dict['id'] = i['id']
                json_dict['golden_money'] = i['golden_money']
                json_dict['integral'] = i['integral']
                data_list.append(json_dict)
            if query_data == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''获取金币(兑换)'''

    def post(self, request):
        try:
            mp = request.POST.get('mp_id')  # 顾问id
            money = request.POST.get('money')  # 要兑换的金币
            if int(money) < 0:
                return JsonResponse({'Result': 400, 'Msg': '输入有误'})
            ig = request.POST.get('integral')  # 要兑换的积分
            if int(ig) < 0:
                return JsonResponse({'Result': 400, 'Msg': '输入有误'})
            req = MiddlePeople.objects.filter(id=mp).values('golden_money', 'integral')
            input_ig = int(ig)
            input_mo = int(money)
            if input_ig / input_mo != 10:
                return JsonResponse({'Result': 400, 'Msg': '金币输入不对'})

            gold = req[0]['golden_money']  # 总金币数
            ing = req[0]['integral']  # 总积分数
            if ing - input_ig < 0:
                return JsonResponse({'Result': 400, 'Msg': '当前总积分数不足以做减法, 请输入积分数足够的数值'})
            else:
                fen = ing - input_ig  # 兑换后剩余积分数
                MiddlePeople.objects.filter(id=mp).update(integral=fen)
                moo = gold + input_mo
                MiddlePeople.objects.filter(id=mp).update(golden_money=moo)
                GoldenMoney.objects.create(fk_id=mp, money_count=input_mo, change_beacuse='兑换金币,数值为%s' % input_mo,
                                           choice_classfiy=0)
                IntegralSubsidiary.objects.create(fk_id=mp, score=input_ig, change_beacuse='积分兑换了金币, 数值为%s' % input_ig,
                                                  choice_classfiy=1)
            data = {'money': fen, 'score': moo}
            return JsonResponse({"statue": '兑换成功', 'data': data}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class GodView(View):
    '''获客神器(实名分享)'''

    def post(self, request):
        try:
            mp = request.POST.get('mp_id')  # 顾问id
            user = request.POST.get('user_id')  # 用户id
            really_name_share = request.POST.get('share_d')  # 实名分享标记
            query_data = ReallyNameMiddlePeopleShare.objects.get(id=mp)
            query_data.user_id = user
            query_data.name = really_name_share
            query_data.middle_id = mp
            query_data.save()
            return JsonResponse({"statue": 'success', 'data': '实名分享修改成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class EveryDayPermitionView(View):
    '''每日任务之签到'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            query_data = MiddlePeople.objects.filter(id=mp).values('integral', 'live_limit')
            query_list = list(query_data)
            data_list = []
            for i in query_list:
                json_dict = {}
                json_dict['integral'] = i['integral']
                json_dict['live_limit'] = i['live_limit']
                data_list.append(json_dict)
            if query_list == []:
                return JsonResponse({'Statue': 'false', 'Msg': 'not_found'})
            return JsonResponse({"statue": 'success', 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def post(self, request):
        '''签到积分修改'''
        try:
            mp = request.POST.get('mp_id')  # 顾问id
            score = request.POST.get('score')  # 积分
            beacuse = '每日签到'
            query_data = MiddlePeople.objects.get(id=mp)
            day = request.POST.get('day')
            scc = MiddlePeople.objects.filter(id=mp).values('integral')
            try:
                sc = scc[0]['integral']
            except Exception as e:
                context = {"Result": 'false', 'Msg': {e}}
                return JsonResponse(context)
            query_data.integral = int(score) + int(sc)
            query_data.save()
            IntegralSubsidiary.objects.create(fk_id=mp, score=5, change_beacuse=beacuse, choice_classfiy=0)
            return JsonResponse({"statue": 'success', 'data': '积分修改成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserCallFreeToMiddlePeopleView(View):
    '''免费拨打电话'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            p = MiddlePeople.objects.get(id=mp)
            phone = p.mobile
            if not p:
                return JsonResponse({'Statue': 'false', 'Msg': '数据获取失败'})
            return JsonResponse({"CODE": '200', 'DATA': phone}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''创建新用户'''

    def post(self, request):
        try:
            mp = request.POST.get('mp_id')  # 顾问id
            user_name = request.POST.get('user_name')  # 用户名
            user_id = request.POST.get('user_id')  # 用户id
            user_phone = request.POST.get('user_phone')  # 用户电话号
            head_img = request.POST.get('head_img')  # 头像
            choice_classfiy = request.POST.get('choice_classfiy')  # 0,1

            MyPhoneCall.objects.create(middle_people_id=mp,
                                       name=user_name,
                                       user_id=user_id,
                                       user_phone=user_phone,
                                       head_img=head_img,
                                       choice_classfiy=choice_classfiy,
                                       )

            return JsonResponse({"CODE": '200', 'MSG': '我的来电数据已添加'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class LoginInTimeGiveValue(View):
    '''判断用户是否是置业顾问'''

    def get(self, request):
        try:
            ud = request.GET.get('user_id')
            p = MiddlePeople.objects.filter(user_fk_id=ud).values('id')
            p_l = list(p)

            if p_l == []:
                return JsonResponse({'Statue': 'false', 'Msg': '该用户不是置业顾问,请引导其进入入驻或浏览界面'})

            return JsonResponse({"CODE": '200', 'DATA': p_l}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''用户登陆状态创建'''

    def post(self, request):
        try:
            ud = request.POST.get('open_id')
            UserSealStatus.objects.create(open_id=ud, statue=False, choice_classfiy=0, if_talk=False)
            return JsonResponse({"CODE": '200', 'DATA': '数据创建成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class PayAttentionToAuthorizedPublicAccountView(View):
    '''关注授权公众号'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 5)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='关注授权公众号', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ShareSmallProgramsWithNewUsersView(View):
    '''分享小程序给新用户'''

    def get(self, request):
        try:
            count = request.GET.get('count')  # 分享次数
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if count == 200:
                return JsonResponse({'Statue': '404', 'Msg': '分享新用户已经达到封顶'})
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 1)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='分享小程序给新用户', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ShareSmallProgramsWithOldUsersView(View):
    '''分享小程序给老用户'''

    def get(self, request):
        try:
            count = request.GET.get('count')  # 分享次数
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if count == 30:
                return JsonResponse({'Statue': '404', 'Msg': '分享新用户已经达到封顶'})
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 1)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='分享小程序给老用户', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class RetweetingMomentsView(View):
    '''转发朋友圈'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 10)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='转发朋友圈', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class IssueShareView(View):
    '''发布分享堂'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 15)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='发布分享唐', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ProvideImportantInformationView(View):
    '''问题反馈'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 10)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='问题反馈', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class PleaseMiddlePeopleView(View):
    '''邀请置业顾问入驻'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + 10)
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse='邀请置业顾问入驻', choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UploadBuildingDetialView(View):
    '''上传楼盘信息'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')  # 顾问id
            score = request.GET.get('score')  # 分数
            choice = request.GET.get('choice')
            sc = MiddlePeople.objects.filter(id=mp).values('integral')
            if sc == []:
                return JsonResponse({'Statue': '404', 'Msg': '数据为空'})
            asc = int(sc[0]['integral'])
            nasc = str(asc + int(score))
            MiddlePeople.objects.filter(id=mp).update(integral=nasc)

            IntegralSubsidiary.objects.create(fk_id=mp, score=nasc, change_beacuse=choice, choice_classfiy=0)

            return JsonResponse({"CODE": '200', 'DATA': '积分添加成功,受影响的表为: 置业顾问, 积分明细表'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserReportView(View):
    '''举报'''

    def post(self, request):
        try:
            up = request.POST.get('user_id')  # 用户id
            classfiy1 = request.POST.get('classfiy1')  # 类型1
            classfiy2 = request.POST.get('classfiy2')  # 类型2
            img_url1 = request.POST.get('img_url')  # 问题图片1
            img_url2 = request.POST.get('img_url')  # 问题图片2
            img_url3 = request.POST.get('img_url')  # 问题图片3
            img_url4 = request.POST.get('img_url')  # 问题图片4
            img_url5 = request.POST.get('img_url')  # 问题图片5
            img_url6 = request.POST.get('img_url')  # 问题图片6
            content = request.POST.get('content')  # 问题描述
            mobile = request.POST.get('mobile')  # 手机号
            user_name = request.POST.get('user_name')  # 用户名
            bad_id = request.POST.get('bad_id')  # 被举报人的id
            UserReport.objects.create(fk_id=up,
                                      classfiy1=classfiy1,
                                      classfiy2=classfiy2,
                                      upload_img1=img_url1,
                                      upload_img2=img_url2,
                                      upload_img3=img_url3,
                                      upload_img4=img_url4,
                                      upload_img5=img_url5,
                                      upload_img6=img_url6,
                                      content=content,
                                      mobile=mobile,
                                      user_name=user_name,
                                      bad_id=bad_id,
                                      )

            return JsonResponse({"STATUE": 'success', 'DATA': '提交成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def get(self, request):  # 文章
        try:
            bd = request.GET.get('at_id')
            up = request.GET.get('user_id')
            req = AttentionAretical.objects.filter(Q(aretical_id=bd) & Q(user_id=up)).count()
            if req == 0:
                return JsonResponse({"statue": 200, 'data': True}, safe=False)
            else:
                return JsonResponse({"statue": 200, 'data': False}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class LotteryRemindView(View):
    '''摇号提醒/判断用户是否关注公众号'''

    def get(self, request):
        try:
            ur = request.GET.get('ur_id')
            bd = request.GET.get('bd_id')
            op_op = Users.objects.get(id=ur)
            try:
                union = MiddleUnionId.objects.get(small_open_id=op_op.open_id)
            except:
                return JsonResponse({"statue": 404, 'data': '不是用户需要关注公众号'})

            return JsonResponse({"statue": 200, 'data': '通知提醒成功!'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''判断用户是否关注了楼盘/如果关注了则默认提醒摇号, 没有则调用关注接口'''

    def post(self, request):  # 楼盘
        try:
            bd = request.POST.get('only_id')
            up = request.POST.get('user_id')
            req = AttentionVillage.objects.filter(Q(building_id=bd) & Q(user_id=up)).count()
            if req == 0:
                return JsonResponse({"statue": 200, 'data': True}, safe=False)  # 没关注
            else:
                return JsonResponse({"statue": 200, 'data': False}, safe=False)  # 关注了
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class IfZanView(View):
    '''判断用户是否关注了顾问/文章'''

    def get(self, request):  # 顾问
        try:
            bd = request.GET.get('middle_people_id')
            up = request.GET.get('user_id')
            req = ExclusiveCustmer.objects.filter(Q(middle_name_id=bd) & Q(user_id=up)).count()
            if req:
                return JsonResponse({"statue": 200, 'data': True}, safe=False)  # 没关注
            else:
                return JsonResponse({"statue": 200, 'data': False}, safe=False)  # 关注了
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    def post(self, request):  # 文章
        try:
            bd = request.POST.get('aretical_id')
            up = request.POST.get('user_id')
            req = AttentionAretical.objects.filter(Q(aretical_id=bd) & Q(user_id=up)).count()
            if req == 0:
                return JsonResponse({"statue": 200, 'data': True}, safe=False)
            else:
                return JsonResponse({"statue": 200, 'data': False}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserSealStatusView(View):
    '''用户封号状态查询'''

    def get(self, request):
        # return HttpResponse(request.GET.get('user_id'))
        try:
            up = request.GET.get('user_id')
            req = UserSealStatus.objects.filter(open_id=up).values()
            req_list = list(req)
            print(req_list)
            data_list = []
            for i in req_list:
                json_dict = {}
                # json_dict['fk_id'] = i['fk_id']
                json_dict['statue'] = i['statue']
                json_dict['choice_classfiy'] = i['choice_classfiy']
                json_dict['if_talk'] = i['if_talk']
                json_dict['create_time'] = i['create_time']
                data_list.append(json_dict)
            return JsonResponse({"statue": 200, 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context, safe=False)

    '''用户聊天记录入库'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            UserHistoryMessageRecord.objects.create(**json_dict)
            return JsonResponse({"statue": 200, 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserChatListView(View):
    '''消息列表页'''

    def get(self, request):
        try:
            up = request.GET.get('ur_id')
            req = MessageChatListView.objects.filter(fk_id=up).values()

            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['fk_id'] = i['fk_id']
                json_dict['id'] = i['id']
                json_dict['create_time'] = i['create_time']
                json_dict['header_image'] = i['header_image']
                # json_dict['message'] = i['message']
                json_dict['name'] = i['name']
                json_dict['mp_id'] = i['pid']
                json_dict['ur_room'] = i['ur_room']
                json_dict['mp_room'] = i['mp_room']
                data_list.append(json_dict)
            return JsonResponse({"statue": 200, 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''顾问聊天记录入库'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            MiddlePeopleHistoryMessageRecord.objects.create(**json_dict)
            return JsonResponse({"statue": 200, 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ExeclUploadView(View):
    '''消息列表页获取最新消息'''

    def get(self, request):
        try:
            up = request.GET.get('ur_id')
            req = UserHistoryMessageRecord.objects.filter(fk_id=up).values()

            req_list = list(req)
            data_list = []
            for i in req_list:
                json_dict = {}
                json_dict['fk_id'] = i['fk_id']
                json_dict['id'] = i['id']
                json_dict['create_time'] = i['create_time']
                json_dict['header_image'] = i['header_image']
                # json_dict['message'] = i['message']
                json_dict['name'] = i['name']
                json_dict['mp_id'] = i['pid']
                json_dict['ur_room'] = i['ur_room']
                json_dict['mp_room'] = i['mp_room']
                data_list.append(json_dict)
            return JsonResponse({"statue": 200, 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class ExeclUploadView(View):
    '''Execl上传一房一价'''

    def post(self, request):
        try:
            hang_number = request.POST.get('hang_number')
            lie_number = request.POST.get('lie_number')
            execl = request.FILES.getlist('execl', None)
            name = request.POST.get('execl_name')
            bd = request.POST.get('bd_id')

            uid = str(uuid.uuid4())
            suid = ''.join(uid.split('-'))
            # name = random.randint(1000000,9999999)
            path_name = str(suid)
            online_path = '/www/wwwroot/www.fangsmalltuan.site/media/{}.xlsx'.format(path_name)

            for f in execl:
                file = open(online_path, 'wb+')
                for chunk in f.chunks():
                    file.write(chunk)
                file.close()

            cname = name.split('#')

            wb = openpyxl.load_workbook(online_path)
            ws = wb.active  # 当前活跃的表单

            for row, cutname in zip(ws[1:int(ws.max_row) - 1], cname):
                # house_dong = str(row[0].value)
                house_yuan = str(row[1].value)
                door_number = str(row[2].value)
                create_area = str(row[3].value)
                in_area = str(row[4].value)
                gave_house = str(row[5].value)
                one_price = str(row[6].value)
                all_price = str(row[7].value)
                # will_sale_number = str(row[8].value)
                public_date = str(row[9].value)
                # give_date = str(row[10].value)
                build_company = str(row[11].value)
                # lottery_title = str(row[12].value)

                house_dong = str(cutname[0])
                all_house_count = str(cutname[1])
                give_date = str(cutname[2])
                house_use = str(cutname[3])
                land_num = str(cutname[4])
                will_sale_number = str(cutname[5])

                OneHouseOnePrice.objects.create(
                    building_detial_id=bd,
                    house_dong=house_dong,
                    house_yuan=house_yuan,
                    door_number=door_number,
                    create_area=create_area,
                    in_area=in_area,
                    gave_house=gave_house,
                    one_price=one_price,
                    all_price=all_price,
                    will_sale_number=will_sale_number,

                    public_date=public_date,
                    give_date=give_date,
                    build_company=build_company,
                    lottery_title=land_num,
                    all_tao_count=all_house_count,
                    house_use=house_use
                )

            os.remove(online_path)
            return JsonResponse({"statue": 200, 'data': '上传成功'}, safe=False)
        except Exception as e:
            context = {'CODE': '400', 'ERROR': {e}}
            return JsonResponse(context)


class UserChatListCreateDelView(View):
    '''消息列表页/删除'''

    def get(self, request):
        try:
            req = request.GET.get('del_id')
            MessageChatListView.objects.filter(id=req).delete()
            return JsonResponse({"statue": 200, 'data': '删除成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''消息列表页/添加'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            MessageChatListView.objects.create(**json_dict)
            return JsonResponse({"statue": 200, 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class UserChatHistoryView(View):
    '''消息详情页'''

    def get(self, request):
        try:
            up = request.GET.get('ur_id')
            mp = request.GET.get('ur_bei_id')
            rn = request.GET.get('room_number')
            req = UserHistoryMessageRecord.objects.filter(room=rn).values().order_by('create_time')
            query_list = list(req)
            data_list = []
            for a in query_list:
                json_dict = {}
                json_dict['id'] = a['id']
                json_dict['fk_id'] = a['fk_id']
                json_dict['user_bei_id'] = a['user_bei_id']
                json_dict['content'] = a['content']
                json_dict['create_time'] = a['create_time']
                json_dict['header_image'] = a['header_image']
                json_dict['room'] = a['room']
                data_list.append(json_dict)
            return JsonResponse({"statue": 200, 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''个人中心任务状态栏创建'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            UserEveryDayPermetionStatue.objects.create(**json_dict)
            return JsonResponse({"statue": 200, 'data': '添加成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)


class EveryDayPermetionView(View):
    '''个人中心任务状态栏显示/加积分-活跃度'''

    def get(self, request):
        try:
            mp = request.GET.get('mp_id')
            import datetime
            date = datetime.datetime.today()
            # req = UserEveryDayPermetionStatue.objects.filter(Q(fk_id=mp)&Q(create_time=date)).values()
            req = UserEveryDayPermetionStatue.objects.filter(fk_id=mp).values()
            mpq = MiddlePeople.objects.filter(id=mp).values('integral', 'live_limit')
            query_list = list(req)
            query_list2 = list(mpq)

            data_list = []
            for a, b in zip(query_list, query_list2):
                json_dict = {}
                # json_dict['integral'] = b['integral']
                # json_dict['live_limit'] = b['live_limit']
                json_dict['id'] = a['id']
                json_dict['fk_id'] = a['fk_id']
                json_dict['q_choice'] = a['q_choice']
                json_dict['q_count'] = a['q_count']
                json_dict['g_choice'] = a['g_choice']
                json_dict['fn_choice'] = a['fn_choice']
                json_dict['fn_count'] = a['fn_count']
                json_dict['fo_choice'] = a['fo_choice']
                json_dict['fo_count'] = a['fo_count']
                json_dict['zf_choice'] = a['zf_choice']
                json_dict['zf_count'] = a['zf_count']
                json_dict['fxt_choice'] = a['fxt_choice']
                json_dict['fxt_count'] = a['fxt_count']
                json_dict['tgxx_choice'] = a['tgxx_choice']
                json_dict['tgxx_count'] = a['tgxx_count']
                json_dict['yqrz_choice'] = a['yqrz_choice']
                json_dict['yqrz_count'] = a['yqrz_count']
                json_dict['sclp_choice'] = a['sclp_choice']
                json_dict['sclp_count'] = a['sclp_count']
                json_dict['create_time'] = a['create_time']
                data_list.append(json_dict)
            return JsonResponse({"statue": 200, 'data': data_list}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)

    '''个人中心任务状态栏/修改状态'''

    def post(self, request):
        try:
            json_dict = json.loads(request.body.decode())
            bd = json_dict.get('fk_id')
            UserEveryDayPermetionStatue.objects.filter(fk_id=bd).update(**json_dict)

            return JsonResponse({"statue": 200, 'data': '修改成功'}, safe=False)
        except Exception as e:
            context = {"Result": 'false', 'Msg': {e}}
            return JsonResponse(context)
